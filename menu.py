import tkinter as tk
from tkinter import ttk
import serial
import serial.tools.list_ports
import time
import openpyxl
import os
from tkinter import filedialog
from tkinter import ttk

# Global variable to keep track of the step number
step_number = 1

def add_selection():
    global step_number

    selected_mode = mode_var.get()
    selected_submode = submode_var.get()
    main_parameter_value = main_parameter_entry.get()

    # Check if inputs are valid
    if selected_mode and (selected_submode or selected_mode == "Rest"):
        if selected_mode != "Rest" and not main_parameter_value:
            return  # Do not add if there's no input for main parameter when required

        # Gather termination parameters
        termination_params = []
        if termination1_var.get() and termination1_entry.get():
            termination_params.append(f"{termination1_var.get()} = {termination1_entry.get()}")
        if termination2_var.get() and termination2_entry.get():
            termination_params.append(f"{termination2_var.get()} = {termination2_entry.get()}")
        if termination3_var.get() and termination3_entry.get():
            termination_params.append(f"{termination3_var.get()} = {termination3_entry.get()}")

        if not termination_params:
            termination_params.append("No Termination Parameters")

        termination_values = " or ".join(termination_params)
        if selected_mode == "Rest":
            main_parameter_value = "Rest"
            main_parameter_unit = ""
        else:
            main_parameter_unit = main_parameter_label.cget("text").split("(")[-1][:-1]  # Extract the unit

        # Check if we are in edit mode
        if edit_mode.get():
            # Edit existing row
            selected_item = table.selection()[0]
            table.item(selected_item, values=(f"Step {table.index(selected_item) + 1}", selected_mode, selected_submode, f"{main_parameter_value} {main_parameter_unit}", termination_values))
            edit_mode.set(False)
            add_button.config(text="Add")
        else:
            # Get selected item
            selected_item = table.selection()
            
            if selected_item:
                selected_index = table.index(selected_item[0])
                
                # Add the new row after the selected row
                new_step_number = step_number
                step_number += 1
                new_row_values = (f"Step {new_step_number}", selected_mode, selected_submode, f"{main_parameter_value} {main_parameter_unit}", termination_values)
                
                # Insert the new row at the position after the selected row
                table.insert('', selected_index + 1, values=new_row_values)
                
                # Update step numbers of all subsequent rows
                for i, item in enumerate(table.get_children()[selected_index + 1:]):
                    table.item(item, values=(f"Step {selected_index + 2 + i}",) + table.item(item, 'values')[1:])
                
            else:
                # Insert into table
                table.insert('', tk.END, values=(f"Step {step_number}", selected_mode, selected_submode, f"{main_parameter_value} {main_parameter_unit}", termination_values))
                step_number += 1  # Increment the step number

        # Clear input fields after adding
        main_parameter_entry.delete(0, tk.END)
        termination1_entry.delete(0, tk.END)
        termination2_entry.delete(0, tk.END)
        termination3_entry.delete(0, tk.END)

        # Reset selection and button text
        table.selection_remove(table.selection())  # Remove selection
        edit_mode.set(False)
        add_button.config(text="Add")

    check_add_button_state()

def update_submenu(mode):
    submodes = []
    if mode == "Charge":
        submodes = ["CC", "CV"]
    elif mode == "DisCharge":
        submodes = ["CC", "CV", "CL", "CP"]
    elif mode == "Rest":
        submodes = ["Rest"]
    
    submode_menu['menu'].delete(0, 'end')  # Clear previous submenu items
    
    for submode in submodes:
        submode_menu['menu'].add_command(label=submode, command=tk._setit(submode_var, submode, update_main_parameter_entry))
    
    submode_var.set(submodes[0])  # Set default selection
    update_main_parameter_entry()
    main_parameter_entry.delete(0, tk.END)  # Reset main parameter entry when menu changes
    check_add_button_state()

def update_main_parameter_entry(event=None):
    selected_mode = mode_var.get()
    selected_submode = submode_var.get()
    
    if selected_mode == "Rest":
        main_parameter_label.config(text="No input required")
        main_parameter_entry.config(state=tk.DISABLED)
    elif selected_mode == "Charge" and selected_submode == "CC":
        main_parameter_label.config(text="Current (mA):")
        main_parameter_entry.config(state=tk.NORMAL)
    elif selected_mode == "Charge" and selected_submode == "CV":
        main_parameter_label.config(text="Voltage (mV):")
        main_parameter_entry.config(state=tk.NORMAL)
    elif selected_mode == "DisCharge" and selected_submode == "CC":
        main_parameter_label.config(text="Current (mA):")
        main_parameter_entry.config(state=tk.NORMAL)
    elif selected_mode == "DisCharge" and selected_submode == "CV":
        main_parameter_label.config(text="Voltage (mV):")
        main_parameter_entry.config(state=tk.NORMAL)
    elif selected_mode == "DisCharge" and selected_submode == "CL":
        main_parameter_label.config(text="Resistance (ohm):")
        main_parameter_entry.config(state=tk.NORMAL)
    elif selected_mode == "DisCharge" and selected_submode == "CP":
        main_parameter_label.config(text="Power (Watt):")
        main_parameter_entry.config(state=tk.NORMAL)
    
    update_termination_parameters(selected_mode, selected_submode)
    termination_frame.grid(row=3, column=0, columnspan=4, padx=5, pady=5)
    check_add_button_state()

def update_termination_parameters(mode, submode):
    # Clear all termination options
    termination1_var.set("")
    termination2_var.set("")
    termination3_var.set("")
    termination1_entry.delete(0, tk.END)
    termination2_entry.delete(0, tk.END)
    termination3_entry.delete(0, tk.END)

    termination_options = []
    if mode == "Charge":
        if submode == "CC":
            termination_options = [("Time", "min"), ("Temp", "°C"), ("Current", "mA")]
        elif submode == "CV":
            termination_options = [("Time", "min"), ("Temp", "°C"), ("Voltage", "mV")]
    elif mode == "DisCharge":
        if submode == "CC":
            termination_options = [("Time", "min"), ("Temp", "°C"), ("Voltage", "mV")]
        elif submode == "CV":
            termination_options = [("Time", "min"), ("Temp", "°C"), ("Current", "mA")]
        elif submode == "CL" or submode == "CP":
            termination_options = [("Time", "min"), ("Temp", "°C"), ("Current", "mA"), ("Voltage", "mV")]
    elif mode == "Rest":
        termination_options = [("Time", "min"), ("Temp", "°C"), ("Current", "mA")]

    if termination_options:
        termination1_label.config(text=termination_options[0][0])
        termination1_unit.config(text=termination_options[0][1])
        termination1_var.set(termination_options[0][0])
        termination1_entry.config(state=tk.NORMAL)
        
        if len(termination_options) > 1:
            termination2_label.config(text=termination_options[1][0])
            termination2_unit.config(text=termination_options[1][1])
            termination2_var.set(termination_options[1][0])
            termination2_entry.config(state=tk.NORMAL)
        else:
            termination2_entry.config(state=tk.DISABLED)
        
        if len(termination_options) > 2:
            termination3_label.config(text=termination_options[2][0])
            termination3_unit.config(text=termination_options[2][1])
            termination3_var.set(termination_options[2][0])
            termination3_entry.config(state=tk.NORMAL)
        else:
            termination3_entry.config(state=tk.DISABLED)

def on_mode_select(mode):
    mode_var.set(mode)
    update_submenu(mode)
    check_add_button_state()

def check_add_button_state(event=None):
    selected_mode = mode_var.get()
    selected_submode = submode_var.get()
    main_parameter_value = main_parameter_entry.get()

    # Check if mode is selected
    if selected_mode and (selected_submode or selected_mode == "Rest"):
        # If mode is Rest, no need to check main parameter value
        if selected_mode == "Rest":
            add_button.config(state=tk.NORMAL)
        else:
            # If mode is not Rest, make sure main parameter value is filled
            if main_parameter_value:
                add_button.config(state=tk.NORMAL)
            else:
                add_button.config(state=tk.DISABLED)
    else:
        add_button.config(state=tk.DISABLED)

def remove_selection():
    global step_number
    selected_item = table.selection()
    if selected_item:
        # Delete the selected item
        table.delete(selected_item)
        
        # Update step numbers
        for i, item in enumerate(table.get_children()):
            table.item(item, values=(f"Step {i+1}",) + table.item(item, 'values')[1:])
        
        # Decrement step_number
        step_number -= 1

def edit_selection():
    selected_item = table.selection()
    if selected_item:
        table.item(selected_item, tags=('selected',))  # Remove selection highlight

        # Retrieve current values
        current_values = table.item(selected_item, 'values')

        # Extract values
        current_mode = current_values[1]
        current_submode = current_values[2]
        main_param_with_unit = current_values[3]
        main_param_value = main_param_with_unit.split()[0] if main_param_with_unit != "No input required" else ""

        # Set the UI with selected values
        mode_var.set(current_mode)
        update_submenu(current_mode)
        submode_var.set(current_submode)
        update_main_parameter_entry()

        if current_mode != "Rest":
            main_parameter_entry.delete(0, tk.END)
            main_parameter_entry.insert(0, main_param_value)

        termination_values = current_values[4].split(" or ")
        for termination_value in termination_values:
            key_value = termination_value.split(" = ")
            if key_value[0] == termination1_var.get():
                termination1_entry.delete(0, tk.END)
                termination1_entry.insert(0, key_value[1])
            elif key_value[0] == termination2_var.get():
                termination2_entry.delete(0, tk.END)
                termination2_entry.insert(0, key_value[1])
            elif key_value[0] == termination3_var.get():
                termination3_entry.delete(0, tk.END)
                termination3_entry.insert(0, key_value[1])

        # Enter edit mode
        edit_mode.set(True)
        add_button.config(text="Save")
        check_add_button_state()

def reset_form():
    global step_number
    mode_var.set("")
    submode_var.set("")
    main_parameter_entry.delete(0, tk.END)
    termination1_entry.delete(0, tk.END)
    termination2_entry.delete(0, tk.END)
    termination3_entry.delete(0, tk.END)
    step_number = 1
    table.delete(*table.get_children())
    check_add_button_state()

def map_mode_submode_param(mode, submode, param, termination_list):
    # Define map for mode
    mode_map = {
        "Charge": "1",
        "DisCharge": "2",
        "Rest": "3"
    }
    
    # Define map for submode
    submode_map = {
        "CC": {"Charge": "1", "DisCharge": "3"},
        "CV": {"Charge": "2", "DisCharge": "4"},
        "CP": "5",
        "CL": "6",
        "Rest": "7"
    }
    
    # Define map for terminations based on the combination of mode and submode
    termination_map = {}
    if mode == "Charge":
        if submode == "CC":
            termination_map = {
                "Voltage": "1",
                "Time": "2",
                "Temp": "3",
            }
        elif submode == "CV":
            termination_map = {
                "Current": "4",
                "Time": "5",
                "Temp": "6",
            }
        elif submode == "CP":
            termination_map = {
                "Voltage": "13",
                "Current": "14",
                "Time": "15",
                "Temp": "16",
            }
    elif mode == "DisCharge":
        if submode == "CC":
            termination_map = {
                "Voltage": "7",
                "Time": "8",
                "Temp": "9",
            }
        elif submode == "CV":
            termination_map = {
                "Current": "10",
                "Time": "11",
                "Temp": "12",
            }
        elif submode == "CP":
            termination_map = {
                "Voltage": "17",
                "Current": "18",
                "Time": "19",
                "Temp": "20",
            }
    elif mode == "Rest":
        termination_map = {
            "Current": "21",
            "Time": "22",
            "Temp": "23",
        }

    # Mapping mode, submode, and terminations
    mode_code = mode_map.get(mode, "0")
    submode_code = submode_map.get(submode, {}).get(mode, submode_map.get(submode, "0"))
    
    termination_codes = [(termination_map.get(term.split(" = ")[0], "0"), term.split(" = ")[1]) for term in termination_list]

    return mode_code, submode_code, termination_codes


def send_to_serial():
    selected_port = port_combobox.get()
    baud_rate = baudrate_var.get()

    if selected_port and baud_rate:
        try:
            ser = serial.Serial(selected_port, baud_rate)
            time.sleep(2)  # Give some time for the connection to establish

            for child in table.get_children():
                values = table.item(child, 'values')
                mode = values[1]
                submode = values[2]
                parameter = values[3].split()[0]  # Take only the numeric value
                terminations = values[4].split(", ")

                # Map the values to numeric codes
                mode_code, submode_code, termination_codes = map_mode_submode_param(mode, submode, parameter, terminations)

                # Construct message
                termination_messages = [f"{code},{value}" for code, value in termination_codes]
                message = f"{mode_code},{submode_code},{parameter}," + ",".join(termination_messages)
                
                # Send the message
                ser.write(message.encode())
                time.sleep(0.5)  # Wait a bit between messages

            ser.close()
        except serial.SerialException as e:
            print(f"Error opening serial port: {e}")



def get_serial_ports():
    ports = serial.tools.list_ports.comports()
    return [port.device for port in ports]


# Function to save data to a text file
def save_to_txt(file_path):
    with open(file_path, 'w') as file:
        for row in table.get_children():
            values = table.item(row, 'values')
            line = '\t'.join(values) + '\n'
            file.write(line)

# Function to open data from a text file
def open_from_txt(file_path):
    with open(file_path, 'r') as file:
        lines = file.readlines()
        table.delete(*table.get_children())  # Clear existing data
        
        for line in lines:
            values = line.strip().split('\t')
            if len(values) == len(table_columns):
                table.insert('', tk.END, values=values)

# Function to save data to an Excel file (excluding header)
def save_to_excel(file_path):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    
    # Write rows (excluding header)
    for row in table.get_children():
        values = table.item(row, 'values')
        sheet.append(values)
    
    workbook.save(file_path)

# Function to open data from an Excel file
def open_from_excel(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    
    table.delete(*table.get_children())  # Clear existing data
    
    for row in sheet.iter_rows(values_only=True):
        if row:  # Make sure the row is not empty
            table.insert('', tk.END, values=row)

# Function to handle save file dialog
def save_file_dialog():
    file_type = [('Text files', '*.txt'), ('Excel files', '*.xlsx')]
    file_path = filedialog.asksaveasfilename(defaultextension=".txt", filetypes=file_type)
    
    if file_path:
        if file_path.endswith('.txt'):
            save_to_txt(file_path)
        elif file_path.endswith('.xlsx'):
            save_to_excel(file_path)

# Function to handle open file dialog
def open_file_dialog():
    file_type = [('Text files', '*.txt'), ('Excel files', '*.xlsx')]
    file_path = filedialog.askopenfilename(filetypes=file_type)
    
    if file_path:
        if file_path.endswith('.txt'):
            open_from_txt(file_path)
        elif file_path.endswith('.xlsx'):
            open_from_excel(file_path)


# Creating the main application window
root = tk.Tk()
root.title("Battery Testing Sequence")


# Create Notebook widget
notebook = ttk.Notebook(root)
notebook.grid(row=0, column=0, sticky="nsew")

# Create first tab for the existing functionality
tab1 = ttk.Frame(notebook)
notebook.add(tab1, text="Main Interface")

# Create second tab for displaying a message
tab2 = ttk.Frame(notebook)
notebook.add(tab2, text="Message")

# Create Menu Bar
menu_bar = tk.Menu(root)
root.config(menu=menu_bar)

# Create File Menu
file_menu = tk.Menu(menu_bar, tearoff=0)
menu_bar.add_cascade(label="File", menu=file_menu)

file_menu.add_command(label="Save", command=lambda: save_file_dialog())
file_menu.add_command(label="Open", command=lambda: open_file_dialog())
# Mode Selection
mode_var = tk.StringVar()
mode_label = ttk.Label(tab1, text="Mode:")
mode_label.grid(row=0, column=0, padx=5, pady=0, sticky="e")
mode_menu = ttk.OptionMenu(tab1, mode_var, "", "Charge", "DisCharge", "Rest", command=on_mode_select)
mode_menu.grid(row=0, column=1, padx=5, pady=5, sticky="w")

# Submode Selection
submode_var = tk.StringVar()
submode_label = ttk.Label(tab1, text="Submode:")
submode_label.grid(row=0, column=2, padx=5, pady=5, sticky="e")
submode_menu = ttk.OptionMenu(tab1, submode_var, "", "Submode")
submode_menu.grid(row=0, column=3, padx=5, pady=5, sticky="w")

# Main Parameter Input
main_parameter_label = ttk.Label(tab1, text="Main Parameter:")
main_parameter_label.grid(row=1, column=0, padx=5, pady=5, sticky="e")
main_parameter_entry = ttk.Entry(tab1)
main_parameter_entry.grid(row=1, column=1, padx=5, pady=5, sticky="w")
main_parameter_entry.bind("<KeyRelease>", check_add_button_state)

# Termination Parameters
termination_frame = ttk.Frame(tab1)
termination1_var = tk.StringVar()
termination1_label = ttk.Label(termination_frame, text="Termination 1:")
termination1_label.grid(row=0, column=0, padx=5, pady=5, sticky="e")
termination1_entry = ttk.Entry(termination_frame)
termination1_entry.grid(row=0, column=1, padx=5, pady=5, sticky="w")
termination1_unit = ttk.Label(termination_frame, text="Unit")
termination1_unit.grid(row=0, column=2, padx=5, pady=5, sticky="w")

termination2_var = tk.StringVar()
termination2_label = ttk.Label(termination_frame, text="Termination 2:")
termination2_label.grid(row=1, column=0, padx=5, pady=5, sticky="e")
termination2_entry = ttk.Entry(termination_frame)
termination2_entry.grid(row=1, column=1, padx=5, pady=5, sticky="w")
termination2_unit = ttk.Label(termination_frame, text="Unit")
termination2_unit.grid(row=1, column=2, padx=5, pady=5, sticky="w")

termination3_var = tk.StringVar()
termination3_label = ttk.Label(termination_frame, text="Termination 3:")
termination3_label.grid(row=2, column=0, padx=5, pady=5, sticky="e")
termination3_entry = ttk.Entry(termination_frame)
termination3_entry.grid(row=2, column=1, padx=5, pady=5, sticky="w")
termination3_unit = ttk.Label(termination_frame, text="Unit")
termination3_unit.grid(row=2, column=2, padx=5, pady=5, sticky="w")

# Add and Reset buttons
button_frame = ttk.Frame(tab1)
button_frame.grid(row=4, column=0, columnspan=4, padx=5, pady=5)

add_button = ttk.Button(button_frame, text="Add", command=add_selection)
add_button.grid(row=0, column=0, padx=5, pady=5)
add_button.config(state=tk.DISABLED)

remove_button = ttk.Button(button_frame, text="Remove", command=remove_selection)
remove_button.grid(row=0, column=1, padx=5, pady=5)

edit_button = ttk.Button(button_frame, text="Edit", command=edit_selection)
edit_button.grid(row=0, column=2, padx=5, pady=5)

reset_button = ttk.Button(button_frame, text="Reset", command=reset_form)
reset_button.grid(row=0, column=3, padx=5, pady=5)

# Edit mode flag
edit_mode = tk.BooleanVar(value=False)

# Sequence Table
table_frame = ttk.Frame(tab1)
table_frame.grid(row=5, column=0, columnspan=4, padx=5, pady=5)

table_columns = ("Step", "Mode", "Submode", "Main Parameter", "Termination Parameters")
table = ttk.Treeview(table_frame, columns=table_columns, show="headings", height=15,)
table.grid(row=0, column=0, sticky="nsew")

for col in table_columns:
    table.heading(col, text=col)
    table.column(col, width=100)

scrollbar = ttk.Scrollbar(table_frame, orient="vertical", command=table.yview)
table.configure(yscroll=scrollbar.set)
scrollbar.grid(row=0, column=1, sticky="ns")

# Serial Communication
serial_frame = ttk.LabelFrame(tab2, text="Serial Communication")
serial_frame.grid(row=6, column=0, columnspan=4, padx=5, pady=5, sticky="ew")

port_label = ttk.Label(serial_frame, text="Port:")
port_label.grid(row=0, column=0, padx=5, pady=5, sticky="e")

port_combobox = ttk.Combobox(serial_frame, values=get_serial_ports())
port_combobox.grid(row=0, column=1, padx=5, pady=5, sticky="w")

baudrate_label = ttk.Label(serial_frame, text="Baudrate:")
baudrate_label.grid(row=0, column=2, padx=5, pady=5, sticky="e")

baudrate_var = tk.StringVar(value="9600")
baudrate_combobox = ttk.Combobox(serial_frame, textvariable=baudrate_var, values=["9600", "19200", "38400", "57600", "115200"])
baudrate_combobox.grid(row=0, column=3, padx=5, pady=5, sticky="w")

send_button = ttk.Button(serial_frame, text="Send", command=send_to_serial)
send_button.grid(row=0, column=4, padx=5, pady=5)

def validate_numeric_input(value):
    # Check if the value is numeric
    return value.isdigit() or value == ""

# Creating the validation command
vcmd = (root.register(validate_numeric_input), '%P')

# Main Parameter Input
main_parameter_entry = ttk.Entry(tab1, validate="key", validatecommand=vcmd)
main_parameter_entry.grid(row=1, column=1, padx=5, pady=5, sticky="w")
main_parameter_entry.bind("<KeyRelease>", check_add_button_state)

# Termination Parameters
termination1_entry = ttk.Entry(termination_frame, validate="key", validatecommand=vcmd)
termination1_entry.grid(row=0, column=1, padx=5, pady=5, sticky="w")
termination2_entry = ttk.Entry(termination_frame, validate="key", validatecommand=vcmd)
termination2_entry.grid(row=1, column=1, padx=5, pady=5, sticky="w")
termination3_entry = ttk.Entry(termination_frame, validate="key", validatecommand=vcmd)
termination3_entry.grid(row=2, column=1, padx=5, pady=5, sticky="w")


# Create a style object
style = ttk.Style()

# Set a theme
style.theme_use("clam")  # You can choose from available themes like 'clam', 'default', 'alt', etc.

# Customize the Treeview widget
style.configure("Treeview",
                background="#f0f0f0",
                foreground="red",
                rowheight=20,
                fieldbackground="#f0f0f0",
                font=('Helvetica', 12))

style.map('Treeview', background=[('selected', '#347083')])

# Customize the headings of the Treeview
style.configure("Treeview.Heading",
                font=('Helvetica', 12, 'bold'),
                background="black",
                foreground="yellow")

# Customize buttons, labels, and entry widgets
style.configure("TButton",
                font=('Helvetica', 12),
                padding=5,
                background="#347083",
                foreground="yellow")

style.configure("TLabel",
                font=('Helvetica', 12),
                padding=1,
                background="#7FFF00",
                foreground="black")

style.configure("TEntry",
                font=('Helvetica', 12),
                padding=5,
                background="#7FFF00",
                foreground="black")

# Sequence Table
table_frame = ttk.Frame(tab1)
table_frame.grid(row=5, column=0, columnspan=4, padx=300, pady=10, sticky="nsew")

table_columns = ("Step", "Mode", "Submode", "Main Parameter", "Termination Parameters")
table = ttk.Treeview(table_frame, columns=table_columns, show="headings", height=20, style="Treeview")
table.grid(row=0, column=0, sticky="nsew")

# Set column headings and adjust column width
table.heading("Step", text="Step")
table.column("Step", anchor=tk.CENTER, width=80)

table.heading("Mode", text="Mode")
table.column("Mode", anchor=tk.CENTER, width=90)

table.heading("Submode", text="Submode")
table.column("Submode", anchor=tk.CENTER, width=100)

table.heading("Main Parameter", text="Main Parameter")
table.column("Main Parameter", anchor=tk.CENTER, width=200)

table.heading("Termination Parameters", text="Termination Parameters")
table.column("Termination Parameters", anchor=tk.CENTER, width=500)

# Add vertical scrollbar
scrollbar = ttk.Scrollbar(table_frame, orient="vertical", command=table.yview)
table.configure(yscroll=scrollbar.set)
scrollbar.grid(row=0, column=1, sticky="ns")

# Add horizontal scrollbar
h_scrollbar = ttk.Scrollbar(table_frame, orient="horizontal", command=table.xview)
table.configure(xscroll=h_scrollbar.set)
h_scrollbar.grid(row=1, column=0, sticky="ew")


root.mainloop()