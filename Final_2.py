import tkinter as tk
from tkinter import ttk
from tkinter import filedialog, messagebox
import serial
import serial.tools.list_ports
import time
import openpyxl
import os
from threading import Thread, Event
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from collections import deque
from datetime import datetime
import webbrowser
from tkcalendar import DateEntry
import pandas as pd
from tkinter import Toplevel, filedialog, messagebox
import matplotlib.pyplot as plt

# Global variables
step_number = 1
final_port = "COM10"
figure = Figure(figsize=(10, 8), dpi=100)
canvas = None
serial_config = {
    "port": None,
    "baud_rate": 9600
}
serial_running = True  # Add a flag to control the serial thread
ser = None
serial_thread = None
connection_event = Event()
running = True
excel_file_path = None
times = deque(maxlen=100)
temperatures = deque(maxlen=100)
voltages = deque(maxlen=100)
currents = deque(maxlen=100)
powers = deque(maxlen=100)
start_time = time.time()
workbook = None
sheet = None

# Function to update the current date and time display
def update_time():
    current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    time_label.config(text=f"Current Date and Time: {current_time}")
    root.after(1000, update_time)  # Update time every second

# Function to open a website in the default web browser
def open_website():
    webbrowser.open("https://www.digikala.com/")  # Replace with the actual URL

def add_selection():
    global step_number

    selected_mode = mode_var.get()
    selected_submode = submode_var.get()
    main_parameter_value = main_parameter_entry.get()

    # Check if inputs are valid
    if selected_mode and (selected_submode or selected_mode == "Rest"):
        if selected_mode != "Rest" and not main_parameter_value:
            return  # Do not add if there's no input for main parameter when required

        # Gather termination parameters
        termination_params = []
        if termination1_var.get() and termination1_entry.get():
            termination_params.append(f"{termination1_var.get()} = {termination1_entry.get()}")
        if termination2_var.get() and termination2_entry.get():
            termination_params.append(f"{termination2_var.get()} = {termination2_entry.get()}")
        if termination3_var.get() and termination3_entry.get():
            termination_params.append(f"{termination3_var.get()} = {termination3_entry.get()}")

        if not termination_params:
            termination_params.append("No Termination Parameters")

        termination_values = " or ".join(termination_params)
        if selected_mode == "Rest":
            main_parameter_value = "Rest"
            main_parameter_unit = ""
        else:
            main_parameter_unit = main_parameter_label.cget("text").split("(")[-1][:-1]  # Extract the unit

        # Check if we are in edit mode
        if edit_mode.get():
            # Edit existing row
            selected_item = table.selection()[0]
            table.item(selected_item, values=(f"Step {table.index(selected_item) + 1}", selected_mode, selected_submode, f"{main_parameter_value} {main_parameter_unit}", termination_values))
            edit_mode.set(False)
            add_button.config(text="Add")
        else:
            # Get selected item
            selected_item = table.selection()
            
            if selected_item:
                selected_index = table.index(selected_item[0])
                
                # Add the new row after the selected row
                new_step_number = step_number
                step_number += 1
                new_row_values = (f"Step {new_step_number}", selected_mode, selected_submode, f"{main_parameter_value} {main_parameter_unit}", termination_values)
                
                # Insert the new row at the position after the selected row
                table.insert('', selected_index + 1, values=new_row_values)
                
                # Update step numbers of all subsequent rows
                for i, item in enumerate(table.get_children()[selected_index + 1:]):
                    table.item(item, values=(f"Step {selected_index + 2 + i}",) + table.item(item, 'values')[1:])
                
            else:
                # Insert into table
                table.insert('', tk.END, values=(f"Step {step_number}", selected_mode, selected_submode, f"{main_parameter_value} {main_parameter_unit}", termination_values))
                step_number += 1  # Increment the step number

        # Clear input fields after adding
        main_parameter_entry.delete(0, tk.END)
        termination1_entry.delete(0, tk.END)
        termination2_entry.delete(0, tk.END)
        termination3_entry.delete(0, tk.END)

        # Reset selection and button text
        table.selection_remove(table.selection())  # Remove selection
        edit_mode.set(False)
        add_button.config(text="Add")

    check_add_button_state()

def update_submenu(mode):
    submodes = []
    if mode == "Charge":
        submodes = ["CC", "CV"]
    elif mode == "DisCharge":
        submodes = ["CC", "CV", "CL", "CP"]
    elif mode == "Rest":
        submodes = ["Rest"]
    
    submode_menu['menu'].delete(0, 'end')  # Clear previous submenu items
    
    for submode in submodes:
        submode_menu['menu'].add_command(label=submode, command=tk._setit(submode_var, submode, update_main_parameter_entry))
    
    submode_var.set(submodes[0])  # Set default selection
    update_main_parameter_entry()
    main_parameter_entry.delete(0, tk.END)  # Reset main parameter entry when menu changes
    check_add_button_state()

def update_main_parameter_entry(event=None):
    selected_mode = mode_var.get()
    selected_submode = submode_var.get()
    
    if selected_mode == "Rest":
        main_parameter_label.config(text="No input required")
        main_parameter_entry.config(state=tk.DISABLED)
    elif selected_mode == "Charge" and selected_submode == "CC":
        main_parameter_label.config(text="Current (mA):")
        main_parameter_entry.config(state=tk.NORMAL)
    elif selected_mode == "Charge" and selected_submode == "CV":
        main_parameter_label.config(text="Voltage (mV):")
        main_parameter_entry.config(state=tk.NORMAL)
    elif selected_mode == "DisCharge" and selected_submode == "CC":
        main_parameter_label.config(text="Current (mA):")
        main_parameter_entry.config(state=tk.NORMAL)
    elif selected_mode == "DisCharge" and selected_submode == "CV":
        main_parameter_label.config(text="Voltage (mV):")
        main_parameter_entry.config(state=tk.NORMAL)
    elif selected_mode == "DisCharge" and selected_submode == "CL":
        main_parameter_label.config(text="Resistance (ohm):")
        main_parameter_entry.config(state=tk.NORMAL)
    elif selected_mode == "DisCharge" and selected_submode == "CP":
        main_parameter_label.config(text="Power (Watt):")
        main_parameter_entry.config(state=tk.NORMAL)
    
    update_termination_parameters(selected_mode, selected_submode)
    termination_frame.grid(row=3, column=0, columnspan=4, padx=5, pady=5)
    check_add_button_state()

def update_termination_parameters(mode, submode):
    # Clear all termination options
    termination1_var.set("")
    termination2_var.set("")
    termination3_var.set("")
    termination1_entry.delete(0, tk.END)
    termination2_entry.delete(0, tk.END)
    termination3_entry.delete(0, tk.END)

    termination_options = []
    if mode == "Charge":
        if submode == "CC":
            termination_options = [("Time", "min"), ("Temp", "°C"), ("Current", "mA")]
        elif submode == "CV":
            termination_options = [("Time", "min"), ("Temp", "°C"), ("Voltage", "mV")]
    elif mode == "DisCharge":
        if submode == "CC":
            termination_options = [("Time", "min"), ("Temp", "°C"), ("Voltage", "mV")]
        elif submode == "CV":
            termination_options = [("Time", "min"), ("Temp", "°C"), ("Current", "mA")]
        elif submode == "CL" or submode == "CP":
            termination_options = [("Time", "min"), ("Temp", "°C"), ("Current", "mA"), ("Voltage", "mV")]
    elif mode == "Rest":
        termination_options = [("Time", "min"), ("Temp", "°C"), ("Current", "mA")]

    if termination_options:
        termination1_label.config(text=termination_options[0][0])
        termination1_unit.config(text=termination_options[0][1])
        termination1_var.set(termination_options[0][0])
        termination1_entry.config(state=tk.NORMAL)
        
        if len(termination_options) > 1:
            termination2_label.config(text=termination_options[1][0])
            termination2_unit.config(text=termination_options[1][1])
            termination2_var.set(termination_options[1][0])
            termination2_entry.config(state=tk.NORMAL)
        else:
            termination2_entry.config(state=tk.DISABLED)
        
        if len(termination_options) > 2:
            termination3_label.config(text=termination_options[2][0])
            termination3_unit.config(text=termination_options[2][1])
            termination3_var.set(termination_options[2][0])
            termination3_entry.config(state=tk.NORMAL)
        else:
            termination3_entry.config(state=tk.DISABLED)

def on_mode_select(mode):
    mode_var.set(mode)
    update_submenu(mode)
    check_add_button_state()

def check_add_button_state(event=None):
    selected_mode = mode_var.get()
    selected_submode = submode_var.get()
    main_parameter_value = main_parameter_entry.get()

    # Check if mode is selected
    if selected_mode and (selected_submode or selected_mode == "Rest"):
        # If mode is Rest, no need to check main parameter value
        if selected_mode == "Rest":
            add_button.config(state=tk.NORMAL)
        else:
            # If mode is not Rest, make sure main parameter value is filled
            if main_parameter_value:
                add_button.config(state=tk.NORMAL)
            else:
                add_button.config(state=tk.DISABLED)
    else:
        add_button.config(state=tk.DISABLED)

def remove_selection():
    global step_number
    selected_item = table.selection()
    if selected_item:
        # Delete the selected item
        table.delete(selected_item)
        
        # Update step numbers
        for i, item in enumerate(table.get_children()):
            table.item(item, values=(f"Step {i+1}",) + table.item(item, 'values')[1:])
        
        # Decrement step_number
        step_number -= 1

def edit_selection():
    selected_item = table.selection()
    if selected_item:
        table.item(selected_item, tags=('selected',))  # Remove selection highlight

        # Retrieve current values
        current_values = table.item(selected_item, 'values')

        # Extract values
        current_mode = current_values[1]
        current_submode = current_values[2]
        main_param_with_unit = current_values[3]
        main_param_value = main_param_with_unit.split()[0] if main_param_with_unit != "No input required" else ""

        # Set the UI with selected values
        mode_var.set(current_mode)
        update_submenu(current_mode)
        submode_var.set(current_submode)
        update_main_parameter_entry()

        if current_mode != "Rest":
            main_parameter_entry.delete(0, tk.END)
            main_parameter_entry.insert(0, main_param_value)

        termination_values = current_values[4].split(" or ")
        for termination_value in termination_values:
            key_value = termination_value.split(" = ")
            if key_value[0] == termination1_var.get():
                termination1_entry.delete(0, tk.END)
                termination1_entry.insert(0, key_value[1])
            elif key_value[0] == termination2_var.get():
                termination2_entry.delete(0, tk.END)
                termination2_entry.insert(0, key_value[1])
            elif key_value[0] == termination3_var.get():
                termination3_entry.delete(0, tk.END)
                termination3_entry.insert(0, key_value[1])

        # Enter edit mode
        edit_mode.set(True)
        add_button.config(text="Save")
        check_add_button_state()

def reset_form():
    global step_number
    mode_var.set("")
    submode_var.set("")
    main_parameter_entry.delete(0, tk.END)
    termination1_entry.delete(0, tk.END)
    termination2_entry.delete(0, tk.END)
    termination3_entry.delete(0, tk.END)
    step_number = 1
    table.delete(*table.get_children())
    check_add_button_state()

def map_mode_submode_param(mode, submode, param, termination_list):
    # Define map for mode
    mode_map = {
        "Charge": "1",
        "DisCharge": "2",
        "Rest": "3"
    }
    
    # Define map for submode
    submode_map = {
        "CC": {"Charge": "1", "DisCharge": "3"},
        "CV": {"Charge": "2", "DisCharge": "4"},
        "CP": "5",
        "CL": "6",
        "Rest": "7"
    }
    
    # Define map for terminations based on the combination of mode and submode
    termination_map = {}
    if mode == "Charge":
        if submode == "CC":
            termination_map = {
                "Voltage": "1",
                "Time": "2",
                "Temp": "3",
            }
        elif submode == "CV":
            termination_map = {
                "Current": "4",
                "Time": "5",
                "Temp": "6",
            }
        elif submode == "CP":
            termination_map = {
                "Voltage": "13",
                "Current": "14",
                "Time": "15",
                "Temp": "16",
            }
    elif mode == "DisCharge":
        if submode == "CC":
            termination_map = {
                "Voltage": "7",
                "Time": "8",
                "Temp": "9",
            }
        elif submode == "CV":
            termination_map = {
                "Current": "10",
                "Time": "11",
                "Temp": "12",
            }
        elif submode == "CP":
            termination_map = {
                "Voltage": "17",
                "Current": "18",
                "Time": "19",
                "Temp": "20",
            }
    elif mode == "Rest":
        termination_map = {
            "Current": "21",
            "Time": "22",
            "Temp": "23",
        }

    # Mapping mode, submode, and terminations
    mode_code = mode_map.get(mode, "0")
    submode_code = submode_map.get(submode, {}).get(mode, submode_map.get(submode, "0"))
    
    termination_codes = [(termination_map.get(term.split(" = ")[0], "0"), term.split(" = ")[1]) for term in termination_list]

    return mode_code, submode_code, termination_codes

# Function to save data to a text file
def save_to_txt(file_path):
    with open(file_path, 'w') as file:
        for row in table.get_children():
            values = table.item(row, 'values')
            line = '\t'.join(values) + '\n'
            file.write(line)

# Function to open data from a text file
def open_from_txt(file_path):
    with open(file_path, 'r') as file:
        lines = file.readlines()
        table.delete(*table.get_children())  # Clear existing data
        
        for line in lines:
            values = line.strip().split('\t')
            if len(values) == len(table_columns):
                table.insert('', tk.END, values=values)

# Function to save data to an Excel file (excluding header)
def save_to_excel(file_path):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    
    # Write rows (excluding header)
    for row in table.get_children():
        values = table.item(row, 'values')
        sheet.append(values)
    
    workbook.save(file_path)

# Function to open data from an Excel file
def open_from_excel(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    
    table.delete(*table.get_children())  # Clear existing data
    
    for row in sheet.iter_rows(values_only=True):
        if row:  # Make sure the row is not empty
            table.insert('', tk.END, values=row)

# Function to handle save file dialog
def save_file_dialog():
    file_type = [('Text files', '*.txt'), ('Excel files', '*.xlsx')]
    file_path = filedialog.asksaveasfilename(defaultextension=".txt", filetypes=file_type)
    
    if file_path:
        if file_path.endswith('.txt'):
            save_to_txt(file_path)
        elif file_path.endswith('.xlsx'):
            save_to_excel(file_path)

# Function to handle open file dialog
def open_file_dialog():
    file_type = [('Text files', '*.txt'), ('Excel files', '*.xlsx')]
    file_path = filedialog.askopenfilename(filetypes=file_type)
    
    if file_path:
        if file_path.endswith('.txt'):
            open_from_txt(file_path)
        elif file_path.endswith('.xlsx'):
            open_from_excel(file_path)

# Function to send data for setting on micro
def send_to_serial():
    global final_port
    selected_port = final_port
    baud_rate = 9600

    if selected_port and baud_rate:
        try:
            ser = serial.Serial(selected_port, baud_rate)
            time.sleep(2)  # Give some time for the connection to establish

            for child in table.get_children():
                values = table.item(child, 'values')
                mode = values[1]
                submode = values[2]
                parameter = values[3].split()[0]  # Take only the numeric value
                terminations = values[4].split(" or ")

                # Map the values to numeric codes
                mode_code, submode_code, termination_codes = map_mode_submode_param(mode, submode, parameter, terminations)

                # Construct message
                termination_messages = [f"{code},{value}" for code, value in termination_codes]
                message = f"{mode_code},{submode_code},{parameter}," + ",".join(termination_messages)

                # Send the message
                ser.write(message.encode())
                time.sleep(0.5)  # Wait a bit between messages

            ser.close()
            messagebox.showinfo("Success", "Data sent successfully")
        except serial.SerialException as e:
            messagebox.showerror("Error", f"Error opening serial port: {e}")

# Functions to identify and prepare the serial port
def find_port():
    global final_port
    final_port = send_messages()
    if final_port:
        messagebox.showinfo("Success", f"Connected to port {final_port}")
    else:
        messagebox.showwarning("Error", "No valid serial port found")
def send_messages():
    # Get the list of serial ports
    ports = get_serial_ports()
    if not ports:
        return None
    for port in ports:
        try:
            ser = serial.Serial(port, baudrate=9600, timeout=1)
            for _ in range(5):
                ser.write(b"HELLO")  # Send "HELLO" message
                time.sleep(1)  # Wait for 1 second
                response = ser.readline().decode().strip()
                if response == "HI":
                    ser.close()
                    return port  # Return the port that responded with "HI"
            ser.close()
        except serial.SerialException:
            continue  # Ignore ports that can't be opened

    return None  # Return None if no port responded with "HI"
def get_serial_ports():
    ports = serial.tools.list_ports.comports()
    return [port.device for port in ports]

#-----------------------------------------PLOT DATA FUNCTIONS---------------------------------------------------
def select_save_path():
    global excel_file_path
    file_path = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        title="Select Save Path"
    )
    if file_path:
        excel_file_path = file_path
        init_excel_file()
        update_status(f"Save path selected: {file_path}")


def init_excel_file():
    global workbook, sheet
    if excel_file_path:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Serial Data"
        sheet.append(["Time (s)", "Seconde", "Temperature", "Voltage", "Current", "Power"])
        workbook.save(excel_file_path)


def log_data_to_excel(time_val, temp, volt, curr, power):
    global workbook, sheet
    if excel_file_path and workbook and sheet:
        absolute_time = start_time + time_val
        formatted_time = datetime.fromtimestamp(absolute_time).strftime("%Y-%m-%d %H:%M:%S")
        elapsed_time_seconds = int(time_val)

        sheet.append([formatted_time, elapsed_time_seconds, temp, volt, curr, power])
        workbook.save(excel_file_path)


def start_serial():
    global serial_thread
    update_status("Connecting...")

    if not serial_config["port"]:
        ports = get_serial_ports()
        if not ports:
            update_status("No serial ports found. Retrying in 5 seconds...")
            root.after(5000, start_serial)
            return
        set_port(final_port)  # Use the final_port value from imported module

    if serial_thread and serial_thread.is_alive():
        update_status("Already connected.")
        return

    serial_thread = Thread(target=serial_reader, args=(serial_config["port"], serial_config["baud_rate"], update_plot))
    serial_thread.start()
    root.after(2000, check_connection)  # Change the interval to 2 seconds


def check_connection():
    if serial_thread and not serial_thread.is_alive():
        update_status("Connection lost. Retrying...")
        root.after(5000, start_serial)
    else:
        root.after(2000, check_connection)  # Change the interval to 2 seconds


def serial_reader(port, baud_rate, callback):
    """
    Read data from the serial port and invoke the callback with the data.

    Args:
    - port (str): The serial port to connect to.
    - baud_rate (int): The baud rate for the serial connection.
    - callback (function): A function to call with the received data.
    """
    global serial_running
    ser = None

    while serial_running:
        try:
            ser = serial.Serial(port, baud_rate, timeout=1)
            callback("Connected")
            while serial_running:
                if ser.in_waiting > 0:
                    try:
                        data = ser.readline()
                        decoded_data = data.decode('utf-8', errors='ignore').strip()
                        callback(decoded_data)
                    except Exception as e:
                        callback(f"Error reading data: {e}")
                time.sleep(2)  # Add a sleep to reduce CPU usage
        except serial.SerialException as e:
            callback(f"Serial exception: {e}")
        except Exception as e:
            callback(f"Unexpected error: {e}")

        callback("Reconnecting...")
        time.sleep(5)


def update_plot(data):
    """
    Update the plot with the new data received from the serial port.

    Args:
    - data (str): The data received from the serial port.
    """
    global times, temperatures, voltages, currents, powers, start_time

    if "Error" in data or "Exception" in data:
        update_status(data)
    elif "Failed to connect" in data:
        update_status("Connection failed. Retrying...")
    else:
        try:
            temp, volt, curr, power = data.split('-')
            current_time = time.time() - start_time

            times.append(current_time)
            temperatures.append(float(temp))
            voltages.append(float(volt))
            currents.append(float(curr))
            powers.append(float(power))

            log_data_to_excel(current_time, temp, volt, curr, power)

            update_plots()

        except ValueError:
            update_status("Received data in unexpected format")


def update_plots():
    global canvas
    figure.clear()

    plot_count = sum(var.get() for var in selected_plots.values())
    if (plot_count == 0):
        canvas.get_tk_widget().grid_forget()
        message_label.grid(row=1, column=0, pady=20, sticky="nsew")
        return

    message_label.grid_forget()
    canvas.get_tk_widget().grid(row=1, column=0, padx=10, pady=10, sticky="nsew")

    rows, cols = (plot_count + 1) // 2, 2
    plot_index = 1

    if selected_plots["Temperature"].get():
        ax = figure.add_subplot(rows, cols, plot_index)
        ax.plot(times, temperatures, label='Temperature', color='r')
        ax.set_title("Temperature", fontsize=12)
        ax.set_xlabel("Time (s)", fontsize=10)
        ax.set_ylabel("Temperature", fontsize=10)
        ax.legend()
        plot_index += 1

    if selected_plots["Voltage"].get():
        ax = figure.add_subplot(rows, cols, plot_index)
        ax.plot(times, voltages, label='Voltage', color='b')
        ax.set_title("Voltage", fontsize=12)
        ax.set_xlabel("Time (s)", fontsize=10)
        ax.set_ylabel("Voltage", fontsize=10)
        ax.legend()
        plot_index += 1

    if selected_plots["Current"].get():
        ax = figure.add_subplot(rows, cols, plot_index)
        ax.plot(times, currents, label='Current', color='g')
        ax.set_title("Current", fontsize=12)
        ax.set_xlabel("Time (s)", fontsize=10)
        ax.set_ylabel("Current", fontsize=10)
        ax.legend()
        plot_index += 1

    if selected_plots["Power"].get():
        ax = figure.add_subplot(rows, cols, plot_index)
        ax.plot(times, powers, label='Power', color='m')
        ax.set_title("Power", fontsize=12)
        ax.set_xlabel("Time (s)", fontsize=10)
        ax.set_ylabel("Power", fontsize=10)
        ax.legend()
        plot_index += 1

    canvas.draw()


def update_status(message):
    """
    Update the status label with the given message.

    Args:
    - message (str): The status message to display.
    """
    if "Connected" in message:
        status_label.config(text=f"Status: {message}", fg="green")
    elif "Error" in message or "Exception" in message:
        status_label.config(text=f"Status: {message}", fg="red")
    else:
        status_label.config(text=f"Status: {message}", fg="orange")


def set_port(port):
    """
    Set the serial port configuration and update the status label.

    Args:
    - port (str): The serial port to set.
    """
    serial_config["port"] = port
    update_status(f"Port set to {port}")


def stop_serial():
    """
    Stop the serial thread if it is running and update the status label.
    """
    global serial_thread, serial_running
    serial_running = False  # Set the flag to stop the thread
    if serial_thread and serial_thread.is_alive():
        serial_thread.join(timeout=5)
    update_status


def close_window():
    root.destroy()  # Close the window immediately

# Creating the main application window
root = tk.Tk()
root.title("Battery Testing Sequence")

# Create Notebook widget
notebook = ttk.Notebook(root)
notebook.grid(row=0, column=0, sticky="nsew")

# Create first tab for the existing functionality
tab1 = ttk.Frame(notebook)
notebook.add(tab1, text="Main Interface")

# Create second tab for displaying a Data
tab2 = ttk.Frame(notebook)
notebook.add(tab2, text="Plot Data")

# Create third tab for plot perives Data
tab3 = ttk.Frame(notebook)
notebook.add(tab3, text="Perives Data")

# Create Menu Bar
menu_bar = tk.Menu(root)
root.config(menu=menu_bar)

# Create File Menu
file_menu = tk.Menu(menu_bar, tearoff=0)
menu_bar.add_cascade(label="File", menu=file_menu)

# Add commands to the File menu with shortcuts
file_menu.add_command(label="Save", accelerator="Ctrl+S", command=lambda: save_file_dialog())
file_menu.add_command(label="Open", accelerator="Ctrl+O", command=lambda: open_file_dialog())
file_menu.add_command(label="Connect", accelerator="Ctrl+C", command=find_port)
file_menu.add_command(label="Save Plot Data", accelerator="Ctrl+P", command=select_save_path)
file_menu.add_command(label="Exit", accelerator="Ctrl+X", command=root.quit)

# Bind shortcuts to the commands
root.bind_all("<Control-s>", lambda event: save_file_dialog())
root.bind_all("<Control-o>", lambda event: open_file_dialog())
root.bind_all("<Control-c>", lambda event: find_port())
root.bind_all("<Control-p>", lambda event: select_save_path())
root.bind_all("<Control-x>", lambda event: root.quit())


# Create About Us Menu
about_menu = tk.Menu(menu_bar, tearoff=0)
menu_bar.add_cascade(label="About Us", menu=about_menu)
about_menu.add_command(label="Visit Our Website", command=open_website)  # Open the website when clicked


# Mode Selection
mode_var = tk.StringVar()
mode_label = ttk.Label(tab1, text="Mode:")
mode_label.grid(row=0, column=0, padx=5, pady=0, sticky="e")
mode_menu = ttk.OptionMenu(tab1, mode_var, "", "Charge", "DisCharge", "Rest", command=on_mode_select)
mode_menu.grid(row=0, column=1, padx=5, pady=5, sticky="w")

# Submode Selection
submode_var = tk.StringVar()
submode_label = ttk.Label(tab1, text="Submode:")
submode_label.grid(row=0, column=2, padx=5, pady=5, sticky="e")
submode_menu = ttk.OptionMenu(tab1, submode_var, "", "Submode")
submode_menu.grid(row=0, column=3, padx=5, pady=5, sticky="w")

# Main Parameter Input
main_parameter_label = ttk.Label(tab1, text="Main Parameter:")
main_parameter_label.grid(row=1, column=0, padx=5, pady=5, sticky="e")
main_parameter_entry = ttk.Entry(tab1)
main_parameter_entry.grid(row=1, column=1, padx=5, pady=5, sticky="w")
main_parameter_entry.bind("<KeyRelease>", check_add_button_state)

# Termination Parameters
termination_frame = ttk.Frame(tab1)
termination1_var = tk.StringVar()
termination1_label = ttk.Label(termination_frame, text="Termination 1:")
termination1_label.grid(row=0, column=0, padx=5, pady=5, sticky="e")
termination1_entry = ttk.Entry(termination_frame)
termination1_entry.grid(row=0, column=1, padx=5, pady=5, sticky="w")
termination1_unit = ttk.Label(termination_frame, text="Unit")
termination1_unit.grid(row=0, column=2, padx=5, pady=5, sticky="w")

termination2_var = tk.StringVar()
termination2_label = ttk.Label(termination_frame, text="Termination 2:")
termination2_label.grid(row=1, column=0, padx=5, pady=5, sticky="e")
termination2_entry = ttk.Entry(termination_frame)
termination2_entry.grid(row=1, column=1, padx=5, pady=5, sticky="w")
termination2_unit = ttk.Label(termination_frame, text="Unit")
termination2_unit.grid(row=1, column=2, padx=5, pady=5, sticky="w")

termination3_var = tk.StringVar()
termination3_label = ttk.Label(termination_frame, text="Termination 3:")
termination3_label.grid(row=2, column=0, padx=5, pady=5, sticky="e")
termination3_entry = ttk.Entry(termination_frame)
termination3_entry.grid(row=2, column=1, padx=5, pady=5, sticky="w")
termination3_unit = ttk.Label(termination_frame, text="Unit")
termination3_unit.grid(row=2, column=2, padx=5, pady=5, sticky="w")

# Add and Reset buttons
button_frame = ttk.Frame(tab1)
button_frame.grid(row=4, column=0, columnspan=4, padx=5, pady=5)

add_button = ttk.Button(button_frame, text="Add", command=add_selection)
add_button.grid(row=0, column=0, padx=5, pady=5)
add_button.config(state=tk.DISABLED)

remove_button = ttk.Button(button_frame, text="Remove", command=remove_selection)
remove_button.grid(row=0, column=1, padx=5, pady=5)

edit_button = ttk.Button(button_frame, text="Edit", command=edit_selection)
edit_button.grid(row=0, column=2, padx=5, pady=5)

reset_button = ttk.Button(button_frame, text="Reset", command=reset_form)
reset_button.grid(row=0, column=3, padx=5, pady=5)

send_button = ttk.Button(button_frame, text="Send", command=send_to_serial)
send_button.grid(row=0, column=4, padx=5, pady=5)

def validate_numeric_input(value):
    # Check if the value is numeric
    return value.isdigit() or value == ""

# Creating the validation command
vcmd = (root.register(validate_numeric_input), '%P')

# Main Parameter Input
main_parameter_entry = ttk.Entry(tab1, validate="key", validatecommand=vcmd)
main_parameter_entry.grid(row=1, column=1, padx=5, pady=5, sticky="w")
main_parameter_entry.bind("<KeyRelease>", check_add_button_state)

# Termination Parameters
termination1_entry = ttk.Entry(termination_frame, validate="key", validatecommand=vcmd)
termination1_entry.grid(row=0, column=1, padx=5, pady=5, sticky="w")
termination2_entry = ttk.Entry(termination_frame, validate="key", validatecommand=vcmd)
termination2_entry.grid(row=1, column=1, padx=5, pady=5, sticky="w")
termination3_entry = ttk.Entry(termination_frame, validate="key", validatecommand=vcmd)
termination3_entry.grid(row=2, column=1, padx=5, pady=5, sticky="w")
# Edit mode flag
edit_mode = tk.BooleanVar(value=False)

# Sequence Table
table_frame = ttk.Frame(tab1)
table_frame.grid(row=5, column=0, columnspan=4, padx=5, pady=5)

table_columns = ("Step", "Mode", "Submode", "Main Parameter", "Termination Parameters")
table = ttk.Treeview(table_frame, columns=table_columns, show="headings", height=15,)
table.grid(row=0, column=0, sticky="nsew")

for col in table_columns:
    table.heading(col, text=col)
    table.column(col, width=100)

scrollbar = ttk.Scrollbar(table_frame, orient="vertical", command=table.yview)
table.configure(yscroll=scrollbar.set)
scrollbar.grid(row=0, column=1, sticky="ns")

# Create a style object
style = ttk.Style()

# Set a theme
style.theme_use("clam")  # You can choose from available themes like 'clam', 'default', 'alt', etc.

# Customize the Treeview widget
style.configure("Treeview",
                background="#f0f0f0",
                foreground="red",
                rowheight=20,
                fieldbackground="#f0f0f0",
                font=('Helvetica', 12))

style.map('Treeview', background=[('selected', '#347083')])

# Customize the headings of the Treeview
style.configure("Treeview.Heading",
                font=('Helvetica', 12, 'bold'),
                background="black",
                foreground="yellow")

# Customize buttons, labels, and entry widgets
style.configure("TButton",
                font=('Helvetica', 12),
                padding=5,
                background="#347083",
                foreground="yellow")

style.configure("TLabel",
                font=('Helvetica', 12),
                padding=1,
                background="#FFD700",
                foreground="black")

style.configure("TEntry",
                font=('Helvetica', 12),
                padding=5,
                background="#FFD700",
                foreground="black")

# Sequence Table
table_frame = ttk.Frame(tab1)
table_frame.grid(row=5, column=0, columnspan=4, padx=300, pady=10, sticky="nsew")

table_columns = ("Step", "Mode", "Submode", "Main Parameter", "Termination Parameters")
table = ttk.Treeview(table_frame, columns=table_columns, show="headings", height=20, style="Treeview")
table.grid(row=0, column=0, sticky="nsew")

# Set column headings and adjust column width
table.heading("Step", text="Step")
table.column("Step", anchor=tk.CENTER, width=80)

table.heading("Mode", text="Mode")
table.column("Mode", anchor=tk.CENTER, width=80)

table.heading("Submode", text="Submode")
table.column("Submode", anchor=tk.CENTER, width=100)

table.heading("Main Parameter", text="Main Parameter")
table.column("Main Parameter", anchor=tk.CENTER, width=200)

table.heading("Termination Parameters", text="Termination Parameters")
table.column("Termination Parameters", anchor=tk.CENTER, width=500)

# Add vertical scrollbar
scrollbar = ttk.Scrollbar(table_frame, orient="vertical", command=table.yview)
table.configure(yscroll=scrollbar.set)
scrollbar.grid(row=0, column=1, sticky="ns")

# Add horizontal scrollbar
h_scrollbar = ttk.Scrollbar(table_frame, orient="horizontal", command=table.xview)
table.configure(xscroll=h_scrollbar.set)
h_scrollbar.grid(row=1, column=0, sticky="ew")

# Plot Data ------------------------------------------------------------------------------------------------------------
root.configure(bg="#f0f0f0")
root.option_add("*Font", "Verdana 10")
root.option_add("*Background", "#f0f0f0")
root.option_add("*Button.Background", "#ffffff")
root.option_add("*Button.Foreground", "#333333")

checkbox_frame = tk.Frame(tab2, bg="#e0e0e0", bd=2, relief=tk.RIDGE)
checkbox_frame.grid(row=0, column=0, pady=10, padx=10, sticky="ew")

selected_plots = {
    "Temperature": tk.BooleanVar(value=False),
    "Voltage": tk.BooleanVar(value=False),
    "Current": tk.BooleanVar(value=False),
    "Power": tk.BooleanVar(value=False)
}

for i, plot_name in enumerate(selected_plots):
    cb = tk.Checkbutton(checkbox_frame, text=plot_name, variable=selected_plots[plot_name],
                        bg="#e0e0e0", fg="#333333", selectcolor="#d0d0d0", padx=10, pady=5)
    cb.grid(row=0, column=i, sticky="w", padx=5)

canvas = FigureCanvasTkAgg(figure, master=tab2)
canvas.get_tk_widget().grid(row=1, column=0, padx=10, pady=10, sticky="nsew")

message_label = tk.Label(tab2, text="Select Parameter To Plot", font=("Verdana", 14, "bold"), fg="#777777", bg="#f0f0f0")
message_label.grid(row=1, column=0, pady=20, sticky="nsew")

status_label = tk.Label(tab2, text="Status: Disconnected", fg="red", bg="#f0f0f0", font=("Verdana", 10, "bold"))
status_label.grid(row=2, column=0, pady=5, padx=10, sticky="ew")

tab2.grid_rowconfigure(1, weight=1)
tab2.grid_columnconfigure(0, weight=1)

start_serial()
set_port(final_port)
root.protocol("WM_DELETE_WINDOW", close_window)

#Perives Dta ploting ---------------------------------------------------------------------------------------------------

# Centering the main frame
main_frame = tk.Frame(tab3, padx=20, pady=20, relief="raised", borderwidth=2)
main_frame.place(relx=0.5, rely=0.5, anchor=tk.CENTER)

# Variable to store the selected file path
file_path = ""


# Function to open a file dialog and select the Excel file
def select_file():
    global file_path
    file_path = filedialog.askopenfilename(
        title="Select Excel File",
        filetypes=[("Excel files", "*.xlsx *.xls")]
    )
    if file_path:
        load_excel_data()


# Function to load the Excel file and create checkboxes for each data column
def load_excel_data():
    global df
    global checkbox_vars
    df = pd.read_excel(file_path)
    columns = df.columns[2:]  # Assuming the first two columns are Time (s) and Seconde

    checkbox_frame = tk.Frame(main_frame, padx=10, pady=10)
    checkbox_frame.grid(row=5, column=0, pady=10, columnspan=2)

    checkbox_vars = []
    for col in columns:
        var = tk.BooleanVar()
        checkbox = tk.Checkbutton(checkbox_frame, text=col, variable=var)
        checkbox.pack(anchor='w')
        checkbox_vars.append((col, var))


# Date picker
tk.Label(main_frame, text="Select Date:", font=("Arial", 12)).grid(row=1, column=0, padx=10, pady=10, sticky='e')
cal = DateEntry(main_frame, width=12, year=2024, month=8, day=31, background='darkgreen', foreground='black',
                borderwidth=2)
cal.grid(row=1, column=1, padx=10, pady=10, sticky='w')

# Hour input
tk.Label(main_frame, text="Select Hour (0-23):", font=("Arial", 12)).grid(row=2, column=0, padx=10, pady=10, sticky='e')
hour_var = tk.StringVar(value="21")
tk.Entry(main_frame, textvariable=hour_var, width=3).grid(row=2, column=1, padx=10, pady=10, sticky='w')

# Button to open file dialog
select_file_button = tk.Button(main_frame, text="Select Excel File", command=select_file, font=("Arial", 12),
                               bg='#4CAF50', fg='white', padx=10, pady=5)
select_file_button.grid(row=0, column=0, padx=10, pady=20, columnspan=2)


def save_plot(fig, col):
    file_path = filedialog.asksaveasfilename(
        defaultextension=".png",
        filetypes=[("PNG files", "*.png"), ("All files", "*.*")]
    )
    if file_path:
        # Save with high resolution
        fig.savefig(file_path, dpi=1000, bbox_inches='tight')
        messagebox.showinfo("Save Plot", f"Plot of {col} saved successfully!")


# Function to plot the data
def plot_data():
    selected_date = cal.get_date()
    selected_hour = hour_var.get()
    selected_datetime = datetime.combine(selected_date, datetime.strptime(selected_hour, "%H").time())

    # Convert the 'Time (s)' column to datetime
    df['Time (s)'] = pd.to_datetime(df['Time (s)'], format='%Y-%m-%d %H:%M:%S')

    # Filter data based on the selected datetime
    filtered_data = df[df['Time (s)'] >= selected_datetime]

    # Select the specific columns where checkboxes are selected
    selected_columns = [col for col, var in checkbox_vars if var.get()]

    if not selected_columns or len(selected_columns) < 1:
        messagebox.showerror("Selection Error", "Please select at least 1 data column.")
        return

    time_series = filtered_data['Seconde']  # 'Seconde' column as the x-axis

    for col in selected_columns:
        # Open a new window for each plot
        plot_window = Toplevel(root)
        plot_window.title(f"Plot of {col}")

        # Adjusted the figure size and added additional parameters
        fig, ax = plt.subplots(
            figsize=(18, 6),  # Make the figure larger
            dpi=100,  # Increase resolution for better quality
            constrained_layout=True,  # Automatically adjust layout to prevent overlap
            subplot_kw={'facecolor': '#ffffff'}  # Set the background color of the subplot
        )

        ax.plot(time_series, filtered_data[col.strip()], label=col.strip())

        ax.set_title(f"Plot of {col} from {selected_datetime}")
        ax.set_xlabel("Seconde")
        ax.set_ylabel(col)
        ax.legend()

        # Display the plot in the new tkinter window
        canvas = FigureCanvasTkAgg(fig, master=plot_window)
        canvas.get_tk_widget().pack()
        canvas.draw()

        # Save button in the new window
        save_button = tk.Button(plot_window, text="Save Plot as Image", command=lambda c=col: save_plot(fig, c),
                                font=("Arial", 10), bg='#4CAF50', fg='white', padx=10, pady=5)
        save_button.pack(pady=10)
# Plot button
plot_button = tk.Button(main_frame, text="Plot Data", command=plot_data, font=("Arial", 12), bg='#4CAF50', fg='white',
                        padx=10, pady=5)
plot_button.grid(row=4, column=0, padx=10, pady=20, columnspan=2)

# Display current date and time
time_label = tk.Label(main_frame, text="", font=("Arial", 12))
time_label.grid(row=6, column=0, pady=10, columnspan=2)
#update_time()


root.mainloop()