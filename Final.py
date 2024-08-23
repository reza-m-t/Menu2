import tkinter as tk
from tkinter import ttk
from tkinter import filedialog, messagebox
import serial
import serial.tools.list_ports
import time
import openpyxl
import os
from threading import Thread, Event
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from collections import deque


# Global variable to keep track of the step number
step_number = 1
final_port = None
ser = None
serial_thread = None
connection_event = Event()
running = True
excel_file_path = None
times = deque(maxlen=100)
temperatures = deque(maxlen=100)
voltages = deque(maxlen=100)
currents = deque(maxlen=100)
powers = deque(maxlen=100)
start_time = time.time()

def connect_to_serial_port(port, baud_rate=9600):
    try:
        ser = serial.Serial(port, baud_rate, timeout=1)
        print(f"Serial port {port} opened with baud rate {baud_rate}.")
        return ser
    except serial.SerialException as e:
        print(f"Error opening serial port {port}: {e}")
        return None

def select_save_path():
    global excel_file_path
    file_path = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        title="Select Save Path"
    )
    if file_path:
        excel_file_path = file_path
        init_excel_file()
        update_status(f"Save path selected: {file_path}")

def init_excel_file():
    if excel_file_path:
        global workbook, sheet
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Serial Data"
        sheet.append(["Time (s)", "Temperature", "Voltage", "Current", "Power"])
        workbook.save(excel_file_path)

def log_data_to_excel(time_val, temp, volt, curr, power):
    if excel_file_path:
        sheet.append([time_val, temp, volt, curr, power])
        workbook.save(excel_file_path)

def start_serial():
    global ser, serial_thread, connection_event
    update_status("Connecting...")
    
    ports = get_serial_ports()
    if not ports:
        update_status("No serial ports found. Retrying in 5 seconds...")
        root.after(5000, start_serial)
        return

    port = final_port
    baud_rate = 9600

    if serial_thread and serial_thread.is_alive():
        update_status("Already connected.")
        return

    serial_thread = Thread(target=serial_read, args=(port, baud_rate, update_plot))
    serial_thread.start()
    root.after(1000, check_connection)

def check_connection():
    if serial_thread and not connection_event.is_set():
        update_status("Connection lost. Retrying...")
        stop_serial()
        root.after(5000, start_serial)
    else:
        root.after(1000, check_connection)

def serial_read(port, baud_rate, callback):
    global ser, running, connection_event
    while running:
        try:
            ser = connect_to_serial_port(port, baud_rate)
            if ser:
                connection_event.set()
                callback("Connected")
                while running:
                    if ser.in_waiting > 0:
                        try:
                            data = ser.readline()
                            decoded_data = data.decode('utf-8', errors='ignore').strip()
                            callback(decoded_data)
                        except Exception as e:
                            callback(f"Error reading data: {e}")
            else:
                callback("Failed to connect")
                connection_event.clear()
        except serial.SerialException as e:
            callback(f"Serial exception: {e}")
        except Exception as e:
            callback(f"Unexpected error: {e}")
        
        callback("Reconnecting...")
        connection_event.clear()
        time.sleep(5)

def stop_serial():
    global ser, running
    running = False
    if ser:
        ser.close()

def update_plot(data):
    global times, temperatures, voltages, currents, powers, start_time
    if "Error reading data" in data or "Serial exception" in data:
        update_status(data)
    elif data == "Failed to connect":
        update_status("Connection failed. Retrying...")
    else:
        try:
            temp, volt, curr, power = data.split('-')
            current_time = time.time() - start_time
            
            times.append(current_time)
            temperatures.append(float(temp))
            voltages.append(float(volt))
            currents.append(float(curr))
            powers.append(float(power))

            log_data_to_excel(current_time, temp, volt, curr, power)

            update_plots()

        except ValueError:
            update_status("Connected")

def update_plots():
    global figure, canvas, selected_plots
    figure.clear()

    plot_count = sum(var.get() for var in selected_plots.values())
    if plot_count == 0:
        canvas.get_tk_widget().pack_forget()
        message_label.place(relx=0.5, rely=0.5, anchor=tk.CENTER)
        return

    message_label.place_forget()
    canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)

    rows, cols = (plot_count + 1) // 2, 2
    plot_index = 1

    if selected_plots["Temperature"].get():
        ax = figure.add_subplot(rows, cols, plot_index)
        ax.plot(times, temperatures, label='Temperature', color='r')
        ax.set_title("Temperature", fontsize=12)
        ax.set_xlabel("Time (s)", fontsize=10)
        ax.set_ylabel("Temperature", fontsize=10)
        ax.legend()
        plot_index += 1

    if selected_plots["Voltage"].get():
        ax = figure.add_subplot(rows, cols, plot_index)
        ax.plot(times, voltages, label='Voltage', color='b')
        ax.set_title("Voltage", fontsize=12)
        ax.set_xlabel("Time (s)", fontsize=10)
        ax.set_ylabel("Voltage", fontsize=10)
        ax.legend()
        plot_index += 1

    if selected_plots["Current"].get():
        ax = figure.add_subplot(rows, cols, plot_index)
        ax.plot(times, currents, label='Current', color='g')
        ax.set_title("Current", fontsize=12)
        ax.set_xlabel("Time (s)", fontsize=10)
        ax.set_ylabel("Current", fontsize=10)
        ax.legend()
        plot_index += 1

    if selected_plots["Power"].get():
        ax = figure.add_subplot(rows, cols, plot_index)
        ax.plot(times, powers, label='Power', color='m')
        ax.set_title("Power", fontsize=12)
        ax.set_xlabel("Time (s)", fontsize=10)
        ax.set_ylabel("Power", fontsize=10)
        ax.legend()
        plot_index += 1

    figure.tight_layout()
    canvas.draw()

def update_status(message):
    status_label.config(text=f"Status: {message}")
    if "Disconnected" in message or "Connection lost" in message or "Connection failed" in message:
        status_label.config(fg="red")
    else:
        status_label.config(fg="green")

def on_close():
    stop_serial()
    if hasattr(globals(), 'workbook') and excel_file_path:
        workbook.close()
    root.destroy()

def add_selection():
    global step_number
    selected_mode = mode_var.get()
    selected_submode = submode_var.get()
    main_parameter_value = main_parameter_entry.get()

    # Check if inputs are valid
    if selected_mode and (selected_submode or selected_mode == "Rest"):
        if selected_mode != "Rest" and not main_parameter_value:
            return  # Do not add if there's no input for main parameter when required
        
        # Gather termination parameters
        termination_params = []
        if termination1_var.get() and termination1_entry.get():
            termination_params.append(f"{termination1_var.get()} = {termination1_entry.get()}")
        if termination2_var.get() and termination2_entry.get():
            termination_params.append(f"{termination2_var.get()} = {termination2_entry.get()}")
        if termination3_var.get() and termination3_entry.get():
            termination_params.append(f"{termination3_var.get()} = {termination3_entry.get()}")

        # Add a default value of 0 if no termination parameters were entered
        if not termination_params:
            termination_params.append("No Termination Parameters")

        termination_values = " or ".join(termination_params)
        if selected_mode == "Rest":
            main_parameter_value = "Rest"
            main_parameter_unit = ""
        else:
            main_parameter_unit = main_parameter_label.cget("text").split("(")[-1][:-1]  # Extract the unit

        if edit_mode.get():
            # Edit existing row
            selected_item = table.selection()[0]
            table.item(selected_item, values=(f"Step {table.index(selected_item) + 1}", selected_mode, selected_submode, f"{main_parameter_value} {main_parameter_unit}", termination_values))
            edit_mode.set(False)
            add_button.config(text="Add")
        else:
            # Insert into table
            table.insert('', tk.END, values=(f"Step {step_number}", selected_mode, selected_submode, f"{main_parameter_value} {main_parameter_unit}", termination_values))
            step_number += 1  # Increment the step number

        # Clear input fields after adding
        main_parameter_entry.delete(0, tk.END)
        termination1_entry.delete(0, tk.END)
        termination2_entry.delete(0, tk.END)
        termination3_entry.delete(0, tk.END)

    check_add_button_state()

def update_submenu(mode):
    submodes = []
    if mode == "Charge":
        submodes = ["CC", "CV"]
    elif mode == "DisCharge":
        submodes = ["CC", "CV", "CL", "CP"]
    elif mode == "Rest":
        submodes = ["Rest"]
    
    submode_menu['menu'].delete(0, 'end')  # Clear previous submenu items
    
    for submode in submodes:
        submode_menu['menu'].add_command(label=submode, command=tk._setit(submode_var, submode, update_main_parameter_entry))
    
    submode_var.set(submodes[0])  # Set default selection
    update_main_parameter_entry()
    main_parameter_entry.delete(0, tk.END)  # Reset main parameter entry when menu changes
    check_add_button_state()

def update_main_parameter_entry(event=None):
    selected_mode = mode_var.get()
    selected_submode = submode_var.get()
    
    if selected_mode == "Rest":
        main_parameter_label.config(text="No input required")
        main_parameter_entry.config(state=tk.DISABLED)
    elif selected_mode == "Charge" and selected_submode == "CC":
        main_parameter_label.config(text="Current (mA):")
        main_parameter_entry.config(state=tk.NORMAL)
    elif selected_mode == "Charge" and selected_submode == "CV":
        main_parameter_label.config(text="Voltage (mV):")
        main_parameter_entry.config(state=tk.NORMAL)
    elif selected_mode == "DisCharge" and selected_submode == "CC":
        main_parameter_label.config(text="Current (mA):")
        main_parameter_entry.config(state=tk.NORMAL)
    elif selected_mode == "DisCharge" and selected_submode == "CV":
        main_parameter_label.config(text="Voltage (mV):")
        main_parameter_entry.config(state=tk.NORMAL)
    elif selected_mode == "DisCharge" and selected_submode == "CL":
        main_parameter_label.config(text="Resistance (ohm):")
        main_parameter_entry.config(state=tk.NORMAL)
    elif selected_mode == "DisCharge" and selected_submode == "CP":
        main_parameter_label.config(text="Power (Watt):")
        main_parameter_entry.config(state=tk.NORMAL)
    
    update_termination_parameters(selected_mode, selected_submode)
    termination_frame.grid(row=3, column=0, columnspan=4, padx=5, pady=5)
    check_add_button_state()

def update_termination_parameters(mode, submode):
    # Clear all termination options
    termination1_var.set("")
    termination2_var.set("")
    termination3_var.set("")
    termination1_entry.delete(0, tk.END)
    termination2_entry.delete(0, tk.END)
    termination3_entry.delete(0, tk.END)

    termination_options = []
    if mode == "Charge":
        if submode == "CC":
            termination_options = [("Time", "min"), ("Temp", "°C"), ("Current", "mA")]
        elif submode == "CV":
            termination_options = [("Time", "min"), ("Temp", "°C"), ("Voltage", "mV")]
    elif mode == "DisCharge":
        if submode == "CC":
            termination_options = [("Time", "min"), ("Temp", "°C"), ("Voltage", "mV")]
        elif submode == "CV":
            termination_options = [("Time", "min"), ("Temp", "°C"), ("Current", "mA")]
        elif submode == "CL" or submode == "CP":
            termination_options = [("Time", "min"), ("Temp", "°C"), ("Current", "mA"), ("Voltage", "mV")]
    elif mode == "Rest":
        termination_options = [("Time", "min"), ("Temp", "°C"), ("Current", "mA")]

    if termination_options:
        termination1_label.config(text=termination_options[0][0])
        termination1_unit.config(text=termination_options[0][1])
        termination1_var.set(termination_options[0][0])
        termination1_entry.config(state=tk.NORMAL)
        
        if len(termination_options) > 1:
            termination2_label.config(text=termination_options[1][0])
            termination2_unit.config(text=termination_options[1][1])
            termination2_var.set(termination_options[1][0])
            termination2_entry.config(state=tk.NORMAL)
        else:
            termination2_entry.config(state=tk.DISABLED)
        
        if len(termination_options) > 2:
            termination3_label.config(text=termination_options[2][0])
            termination3_unit.config(text=termination_options[2][1])
            termination3_var.set(termination_options[2][0])
            termination3_entry.config(state=tk.NORMAL)
        else:
            termination3_entry.config(state=tk.DISABLED)

def on_mode_select(mode):
    mode_var.set(mode)
    update_submenu(mode)
    check_add_button_state()

def check_add_button_state(event=None):
    selected_mode = mode_var.get()
    selected_submode = submode_var.get()
    main_parameter_value = main_parameter_entry.get()

    # Check if mode is selected
    if selected_mode and (selected_submode or selected_mode == "Rest"):
        # If mode is Rest, no need to check main parameter value
        if selected_mode == "Rest":
            add_button.config(state=tk.NORMAL)
        else:
            # If mode is not Rest, make sure main parameter value is filled
            if main_parameter_value:
                add_button.config(state=tk.NORMAL)
            else:
                add_button.config(state=tk.DISABLED)
    else:
        add_button.config(state=tk.DISABLED)

def remove_selection():
    global step_number
    selected_item = table.selection()
    if selected_item:
        # Delete the selected item
        table.delete(selected_item)
        
        # Update step numbers
        for i, item in enumerate(table.get_children()):
            table.item(item, values=(f"Step {i+1}",) + table.item(item, 'values')[1:])
        
        # Decrement step_number
        step_number -= 1

def edit_selection():
    selected_item = table.selection()
    if selected_item:
        table.item(selected_item, tags=('selected',))  # Remove selection highlight

        # Retrieve current values
        current_values = table.item(selected_item, 'values')

        # Extract values
        current_mode = current_values[1]
        current_submode = current_values[2]
        main_param_with_unit = current_values[3]
        main_param_value = main_param_with_unit.split()[0] if main_param_with_unit != "No input required" else ""

        # Set the UI with selected values
        mode_var.set(current_mode)
        update_submenu(current_mode)
        submode_var.set(current_submode)
        update_main_parameter_entry()

        if current_mode != "Rest":
            main_parameter_entry.delete(0, tk.END)
            main_parameter_entry.insert(0, main_param_value)

        termination_values = current_values[4].split(" or ")
        for termination_value in termination_values:
            key_value = termination_value.split(" = ")
            if key_value[0] == termination1_var.get():
                termination1_entry.delete(0, tk.END)
                termination1_entry.insert(0, key_value[1])
            elif key_value[0] == termination2_var.get():
                termination2_entry.delete(0, tk.END)
                termination2_entry.insert(0, key_value[1])
            elif key_value[0] == termination3_var.get():
                termination3_entry.delete(0, tk.END)
                termination3_entry.insert(0, key_value[1])

        # Enter edit mode
        edit_mode.set(True)
        add_button.config(text="Save")
        check_add_button_state()

def reset_form():
    global step_number
    mode_var.set("")
    submode_var.set("")
    main_parameter_entry.delete(0, tk.END)
    termination1_entry.delete(0, tk.END)
    termination2_entry.delete(0, tk.END)
    termination3_entry.delete(0, tk.END)
    step_number = 1
    table.delete(*table.get_children())
    check_add_button_state()

def map_mode_submode_param(mode, submode, param, termination_list):
    # Define map for mode
    mode_map = {
        "Charge": "1",
        "DisCharge": "2",
        "Rest": "3"
    }
    
    # Define map for submode
    submode_map = {
        "CC": {"Charge": "1", "DisCharge": "3"},
        "CV": {"Charge": "2", "DisCharge": "4"},
        "CP": "5",
        "CL": "6",
        "Rest": "7"
    }
    
    # Define map for terminations based on the combination of mode and submode
    termination_map = {}
    if mode == "Charge":
        if submode == "CC":
            termination_map = {
                "Voltage": "1",
                "Time": "2",
                "Temp": "3",
            }
        elif submode == "CV":
            termination_map = {
                "Current": "4",
                "Time": "5",
                "Temp": "6",
            }
        elif submode == "CP":
            termination_map = {
                "Voltage": "13",
                "Current": "14",
                "Time": "15",
                "Temp": "16",
            }
    elif mode == "DisCharge":
        if submode == "CC":
            termination_map = {
                "Voltage": "7",
                "Time": "8",
                "Temp": "9",
            }
        elif submode == "CV":
            termination_map = {
                "Current": "10",
                "Time": "11",
                "Temp": "12",
            }
        elif submode == "CP":
            termination_map = {
                "Voltage": "17",
                "Current": "18",
                "Time": "19",
                "Temp": "20",
            }
    elif mode == "Rest":
        termination_map = {
            "Current": "21",
            "Time": "22",
            "Temp": "23",
        }

    # Mapping mode, submode, and terminations
    mode_code = mode_map.get(mode, "0")
    submode_code = submode_map.get(submode, {}).get(mode, submode_map.get(submode, "0"))
    
    termination_codes = [(termination_map.get(term.split(" = ")[0], "0"), term.split(" = ")[1]) for term in termination_list]

    return mode_code, submode_code, termination_codes

def find_port():
    global final_port
    final_port = send_messages()
    if final_port:
        messagebox.showinfo("Success", f"Connected to port {final_port}")
    else:
        messagebox.showwarning("Error", "No valid serial port found")

def send_messages():
    # Get the list of serial ports
    ports = get_serial_ports()
    if not ports:
        return None

    for port in ports:
        try:
            ser = serial.Serial(port, baudrate=9600, timeout=1)
            for _ in range(5):
                ser.write(b"HELLO")  # Send "HELLO" message
                time.sleep(1)  # Wait for 1 second
                response = ser.readline().decode().strip()
                if response == "HI":
                    ser.close()
                    return port  # Return the port that responded with "HI"
            ser.close()
        except serial.SerialException:
            continue  # Ignore ports that can't be opened
    
    return None  # Return None if no port responded with "HI"

def send_to_serial():
    global final_port
    selected_port = final_port
    baud_rate = 9600

    if selected_port and baud_rate:
        try:
            ser = serial.Serial(selected_port, baud_rate)
            time.sleep(2)  # Give some time for the connection to establish

            for child in table.get_children():
                values = table.item(child, 'values')
                mode = values[1]
                submode = values[2]
                parameter = values[3].split()[0]  # Take only the numeric value
                terminations = values[4].split(" or ")

                # Map the values to numeric codes
                mode_code, submode_code, termination_codes = map_mode_submode_param(mode, submode, parameter, terminations)

                # Construct message
                termination_messages = [f"{code},{value}" for code, value in termination_codes]
                message = f"{mode_code},{submode_code},{parameter}," + ",".join(termination_messages)
                
                # Send the message
                ser.write(message.encode())
                time.sleep(0.5)  # Wait a bit between messages

            ser.close()
            messagebox.showinfo("Success", "Data sent successfully")
        except serial.SerialException as e:
            messagebox.showerror("Error", f"Error opening serial port: {e}")

def get_serial_ports():
    ports = serial.tools.list_ports.comports()
    return [port.device for port in ports]

# Function to save data to a text file
def save_to_txt(file_path):
    with open(file_path, 'w') as file:
        for row in table.get_children():
            values = table.item(row, 'values')
            line = '\t'.join(values) + '\n'
            file.write(line)

# Function to open data from a text file
def open_from_txt(file_path):
    with open(file_path, 'r') as file:
        lines = file.readlines()
        table.delete(*table.get_children())  # Clear existing data
        
        for line in lines:
            values = line.strip().split('\t')
            if len(values) == len(table_columns):
                table.insert('', tk.END, values=values)

# Function to save data to an Excel file (excluding header)
def save_to_excel(file_path):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    
    # Write rows (excluding header)
    for row in table.get_children():
        values = table.item(row, 'values')
        sheet.append(values)
    
    workbook.save(file_path)

# Function to open data from an Excel file
def open_from_excel(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    
    table.delete(*table.get_children())  # Clear existing data
    
    for row in sheet.iter_rows(values_only=True):
        if row:  # Make sure the row is not empty
            table.insert('', tk.END, values=row)

# Function to handle save file dialog
def save_file_dialog():
    file_type = [('Text files', '*.txt'), ('Excel files', '*.xlsx')]
    file_path = filedialog.asksaveasfilename(defaultextension=".txt", filetypes=file_type)
    
    if file_path:
        if file_path.endswith('.txt'):
            save_to_txt(file_path)
        elif file_path.endswith('.xlsx'):
            save_to_excel(file_path)

# Function to handle open file dialog
def open_file_dialog():
    file_type = [('Text files', '*.txt'), ('Excel files', '*.xlsx')]
    file_path = filedialog.askopenfilename(filetypes=file_type)
    
    if file_path:
        if file_path.endswith('.txt'):
            open_from_txt(file_path)
        elif file_path.endswith('.xlsx'):
            open_from_excel(file_path)

# Creating the main application window
root = tk.Tk()
root.title("Battery Testing Sequence")

# Create Notebook widget
notebook = ttk.Notebook(root)
notebook.grid(row=0, column=0, sticky="nsew")

# Create first tab for the existing functionality
tab1 = ttk.Frame(notebook)
notebook.add(tab1, text="Main Interface")

# Create second tab for displaying a message
tab2 = ttk.Frame(notebook)
notebook.add(tab2, text="Plot Data")

# Create Menu Bar
menu_bar = tk.Menu(root)
root.config(menu=menu_bar)

# Create File Menu
file_menu = tk.Menu(menu_bar, tearoff=0)
menu_bar.add_cascade(label="File", menu=file_menu)

file_menu.add_command(label="Save", command=lambda: save_file_dialog())
file_menu.add_command(label="Open", command=lambda: open_file_dialog())
file_menu.add_command(label="Connect", command=find_port)
file_menu.add_command(label="Exit", command=root.quit)
file_menu.add_command(label="Select Save Path", command=select_save_path)

# Mode Selection
mode_var = tk.StringVar()
mode_label = ttk.Label(tab1, text="Mode:")
mode_label.grid(row=0, column=0, padx=5, pady=0, sticky="e")
mode_menu = ttk.OptionMenu(tab1, mode_var, "", "Charge", "DisCharge", "Rest", command=on_mode_select)
mode_menu.grid(row=0, column=1, padx=5, pady=5, sticky="w")

# Submode Selection
submode_var = tk.StringVar()
submode_label = ttk.Label(tab1, text="Submode:")
submode_label.grid(row=0, column=2, padx=5, pady=5, sticky="e")
submode_menu = ttk.OptionMenu(tab1, submode_var, "", "Submode")
submode_menu.grid(row=0, column=3, padx=5, pady=5, sticky="w")

# Main Parameter Input
main_parameter_label = ttk.Label(tab1, text="Main Parameter:")
main_parameter_label.grid(row=1, column=0, padx=5, pady=5, sticky="e")
main_parameter_entry = ttk.Entry(tab1)
main_parameter_entry.grid(row=1, column=1, padx=5, pady=5, sticky="w")
main_parameter_entry.bind("<KeyRelease>", check_add_button_state)

# Termination Parameters
termination_frame = ttk.Frame(tab1)
termination1_var = tk.StringVar()
termination1_label = ttk.Label(termination_frame, text="Termination 1:")
termination1_label.grid(row=0, column=0, padx=5, pady=5, sticky="e")
termination1_entry = ttk.Entry(termination_frame)
termination1_entry.grid(row=0, column=1, padx=5, pady=5, sticky="w")
termination1_unit = ttk.Label(termination_frame, text="Unit")
termination1_unit.grid(row=0, column=2, padx=5, pady=5, sticky="w")

termination2_var = tk.StringVar()
termination2_label = ttk.Label(termination_frame, text="Termination 2:")
termination2_label.grid(row=1, column=0, padx=5, pady=5, sticky="e")
termination2_entry = ttk.Entry(termination_frame)
termination2_entry.grid(row=1, column=1, padx=5, pady=5, sticky="w")
termination2_unit = ttk.Label(termination_frame, text="Unit")
termination2_unit.grid(row=1, column=2, padx=5, pady=5, sticky="w")

termination3_var = tk.StringVar()
termination3_label = ttk.Label(termination_frame, text="Termination 3:")
termination3_label.grid(row=2, column=0, padx=5, pady=5, sticky="e")
termination3_entry = ttk.Entry(termination_frame)
termination3_entry.grid(row=2, column=1, padx=5, pady=5, sticky="w")
termination3_unit = ttk.Label(termination_frame, text="Unit")
termination3_unit.grid(row=2, column=2, padx=5, pady=5, sticky="w")

# Add and Reset buttons
button_frame = ttk.Frame(tab1)
button_frame.grid(row=4, column=0, columnspan=4, padx=5, pady=5)

add_button = ttk.Button(button_frame, text="Add", command=add_selection)
add_button.grid(row=0, column=0, padx=5, pady=5)
add_button.config(state=tk.DISABLED)

remove_button = ttk.Button(button_frame, text="Remove", command=remove_selection)
remove_button.grid(row=0, column=1, padx=5, pady=5)

edit_button = ttk.Button(button_frame, text="Edit", command=edit_selection)
edit_button.grid(row=0, column=2, padx=5, pady=5)

reset_button = ttk.Button(button_frame, text="Reset", command=reset_form)
reset_button.grid(row=0, column=3, padx=5, pady=5)

send_button = ttk.Button(button_frame, text="Send", command=send_to_serial)
send_button.grid(row=0, column=4, padx=5, pady=5)

# Edit mode flag
edit_mode = tk.BooleanVar(value=False)

# Sequence Table
table_frame = ttk.Frame(tab1)
table_frame.grid(row=5, column=0, columnspan=4, padx=5, pady=5)

table_columns = ("Step", "Mode", "Submode", "Main Parameter", "Termination Parameters")
table = ttk.Treeview(table_frame, columns=table_columns, show="headings", height=15,)
table.grid(row=0, column=0, sticky="nsew")

for col in table_columns:
    table.heading(col, text=col)
    table.column(col, width=100)

scrollbar = ttk.Scrollbar(table_frame, orient="vertical", command=table.yview)
table.configure(yscroll=scrollbar.set)
scrollbar.grid(row=0, column=1, sticky="ns")

# Create a style object
style = ttk.Style()

# Set a theme
style.theme_use("clam")  # You can choose from available themes like 'clam', 'default', 'alt', etc.

# Customize the Treeview widget
style.configure("Treeview",
                background="#f0f0f0",
                foreground="red",
                rowheight=20,
                fieldbackground="#f0f0f0",
                font=('Helvetica', 12))

style.map('Treeview', background=[('selected', '#347083')])

# Customize the headings of the Treeview
style.configure("Treeview.Heading",
                font=('Helvetica', 12, 'bold'),
                background="black",
                foreground="yellow")

# Customize buttons, labels, and entry widgets
style.configure("TButton",
                font=('Helvetica', 12),
                padding=5,
                background="#347083",
                foreground="yellow")

style.configure("TLabel",
                font=('Helvetica', 12),
                padding=1,
                background="#FFD700",
                foreground="black")

style.configure("TEntry",
                font=('Helvetica', 12),
                padding=5,
                background="#FFD700",
                foreground="black")

# Sequence Table
table_frame = ttk.Frame(tab1)
table_frame.grid(row=5, column=0, columnspan=4, padx=300, pady=10, sticky="nsew")

table_columns = ("Step", "Mode", "Submode", "Main Parameter", "Termination Parameters")
table = ttk.Treeview(table_frame, columns=table_columns, show="headings", height=20, style="Treeview")
table.grid(row=0, column=0, sticky="nsew")

# Set column headings and adjust column width
table.heading("Step", text="Step")
table.column("Step", anchor=tk.CENTER, width=80)

table.heading("Mode", text="Mode")
table.column("Mode", anchor=tk.CENTER, width=80)

table.heading("Submode", text="Submode")
table.column("Submode", anchor=tk.CENTER, width=100)

table.heading("Main Parameter", text="Main Parameter")
table.column("Main Parameter", anchor=tk.CENTER, width=200)

table.heading("Termination Parameters", text="Termination Parameters")
table.column("Termination Parameters", anchor=tk.CENTER, width=500)

# Add vertical scrollbar
scrollbar = ttk.Scrollbar(table_frame, orient="vertical", command=table.yview)
table.configure(yscroll=scrollbar.set)
scrollbar.grid(row=0, column=1, sticky="ns")

# Add horizontal scrollbar
h_scrollbar = ttk.Scrollbar(table_frame, orient="horizontal", command=table.xview)
table.configure(xscroll=h_scrollbar.set)
h_scrollbar.grid(row=1, column=0, sticky="ew")


root.configure(bg="#f0f0f0")
root.option_add("*Font", "Verdana 10")
root.option_add("*Background", "#f0f0f0")
root.option_add("*Button.Background", "#ffffff")
root.option_add("*Button.Foreground", "#333333")

checkbox_frame = tk.Frame(tab2, bg="#e0e0e0", bd=2, relief=tk.RIDGE)
checkbox_frame.pack(pady=10, padx=10, fill=tk.X)

selected_plots = {
    "Temperature": tk.BooleanVar(value=False),
    "Voltage": tk.BooleanVar(value=False),
    "Current": tk.BooleanVar(value=False),
    "Power": tk.BooleanVar(value=False)
}

for plot_name in selected_plots:
    cb = tk.Checkbutton(checkbox_frame, text=plot_name, variable=selected_plots[plot_name], 
                        bg="#e0e0e0", fg="#333333", selectcolor="#d0d0d0", padx=10, pady=5)
    cb.pack(side=tk.LEFT, anchor="w", padx=5)

figure = Figure(figsize=(10, 8), dpi=100)
canvas = FigureCanvasTkAgg(figure, master=tab2)
canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)

message_label = tk.Label(tab2, text="Select Parameter To Plot", font=("Verdana", 14, "bold"), fg="#777777", bg="#f0f0f0")
message_label.place(relx=0.5, rely=0.5, anchor=tk.CENTER)

status_label = tk.Label(tab2, text="Status: Disconnected", fg="red", bg="#f0f0f0", font=("Verdana", 10, "bold"))
status_label.pack(pady=5, padx=10)

root.mainloop()